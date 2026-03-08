from django.contrib import admin
from django.http import HttpResponse
from .models import EventSettings, Speaker, Schedule, Registration
import openpyxl
from openpyxl.styles import Font
from django.utils import timezone

def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=registrations_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Регистрации'

    # Заголовки
    headers = ['ID', 'Имя', 'Email', 'Телефон', 'Дата регистрации']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Данные
    for row_num, reg in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1).value = reg.id
        ws.cell(row=row_num, column=2).value = reg.name
        ws.cell(row=row_num, column=3).value = reg.email
        ws.cell(row=row_num, column=4).value = reg.phone
        ws.cell(row=row_num, column=5).value = reg.registered_at.strftime('%d.%m.%Y %H:%M')

    wb.save(response)
    return response

export_to_excel.short_description = "Выгрузить выбранные регистрации в Excel"

@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'phone', 'registered_at')
    search_fields = ('name', 'email')
    actions = [export_to_excel]

@admin.register(Speaker)
class SpeakerAdmin(admin.ModelAdmin):
    list_display = ('name', 'title')
    search_fields = ('name',)

@admin.register(Schedule)
class ScheduleAdmin(admin.ModelAdmin):
    list_display = ('start_time', 'title', 'speaker')
    list_filter = ('start_time',)

@admin.register(EventSettings)
class EventSettingsAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        return False if EventSettings.objects.exists() else True